import datetime
from configparser import ConfigParser
from datetime import date
from openpyxl import Workbook, load_workbook
from dateutil.relativedelta import relativedelta, MO


class ReportAutomation():

	def config_setup(self):
		file = 'config.ini'
		self.config = ConfigParser()
		self.config.read(file)


	def generate_dates(self):
		day = datetime.timedelta(1)

		self.mon = date.today() + relativedelta(weekday=MO(0))
		self.tue = self.mon + day
		self.wed = self.mon + day*2
		self.thur = self.mon + day*3
		self.fri = self.mon + day*4
		self.sat = self.mon + day*5
		self.sun = self.mon + day*6


	def format_dates(self):
		# Format: "26-Apr-22"
		self.mon_date = self.mon.strftime("%d-%b-%y") 
		self.tue_date = self.tue.strftime("%d-%b-%y")
		self.wed_date = self.wed.strftime("%d-%b-%y")
		self.thur_date = self.thur.strftime("%d-%b-%y")
		self.fri_date = self.fri.strftime("%d-%b-%y")
		self.sat_date = self.sat.strftime("%d-%b-%y")
		self.sun_date = self.sun.strftime("%d-%b-%y")
		# Format: "April 26 - April 27"
		self.mon_wkst_title_date = self.mon.strftime("%B %d")
		self.fri_wkst_title_date = self.fri.strftime("%B %d")
		# Format: "4/26/22 - 4/27/22"
		self.mon_title_date = self.mon.strftime("%#m/%#d/%y")
		self.fri_title_date = self.fri.strftime("%#m/%#d/%y")


	def setup_excel(self):
		self.wb = load_workbook(self.config['location']['address'])
		self.wb.active = self.wb['TEMPLATE']
		self.ws = self.wb.active
		target = self.wb.copy_worksheet(self.ws)
		target.sheet_view.zoomScale = 80
		target.title = f"{self.mon_wkst_title_date} - {self.fri_wkst_title_date}"
		self.ws = self.wb[f"{self.mon_wkst_title_date} - {self.fri_wkst_title_date}"]


	def populate_cells(self):
		self.ws['A1'].value = f"Weekly Report   ({self.mon_title_date} - {self.fri_title_date})"
		self.ws['B2'].value = self.mon_date
		self.ws['D2'].value = self.tue_date
		self.ws['F2'].value = self.wed_date
		self.ws['H2'].value = self.thur_date
		self.ws['J2'].value = self.fri_date
		self.ws['L2'].value = self.sat_date
		self.ws['N2'].value = self.sun_date


	def save_workbook(self):
		self.wb.save(self.config['location']['address'])