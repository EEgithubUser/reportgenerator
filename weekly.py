# this program will run every week starting Monday (Using Microsoft Tasks) to help configure datetime
import datetime
from configparser import ConfigParser
from datetime import date
from openpyxl import Workbook, load_workbook

file = 'config.ini'
config = ConfigParser()
config.read(file)

# Getting the date
dt = datetime.timedelta(1)

mon = date.today() # Generate today's date
tue = mon + dt
wed = mon + dt*2
thur = mon + dt*3
fri = mon + dt*4
sat = mon + dt*5
sun = mon + dt*6

# Convert to string format in d-mon-y
first = mon.strftime("%d-%b-%y") # string format for today in m/d/y
second = tue.strftime("%d-%b-%y")
third = wed.strftime("%d-%b-%y")
fourth = thur.strftime("%d-%b-%y")
fifth = fri.strftime("%d-%b-%y")
sixth = sat.strftime("%d-%b-%y")
seventh = sun.strftime("%d-%b-%y")

wb = load_workbook(config['location']['address'])

# Copy Weekly Report Template and create worksheet
wb.active = wb['TEMPLATE'] # worksheet that is currently ACTIVE!
ws = wb.active
target = wb.copy_worksheet(ws) # Where "ws" is the source/template

begin = mon.strftime("%B %d") # Format Monday to (Month Full Name Day of Month)
end = fri.strftime("%B %d") # Format Friday to (Month Full Name Day of Month)
target.title = f"{begin} - {end}" # Change name of created worksheet
target.sheet_view.zoomScale = 80 # Set zoom to 80% for the worksheet we just created
ws = wb[f"{begin} - {end}"] # Worksheet in which we are editing is now the one we created in the line above

# "4/26/22 - 4/27/22"
# Use # after % to remove leading 0s 
begin_title = mon.strftime("%#m/%#d/%y")
end_title = fri.strftime("%#m/%#d/%y")
# Pull date and change title to "Weekly Report (Date - Date)"
ws['A1'].value = f"Weekly Report   ({begin_title} - {end_title})"

# Pull date and change Cell B2 - Cell N2 accordingly
ws['B2'].value = first
ws['D2'].value = second
ws['F2'].value = third
ws['H2'].value = fourth
ws['J2'].value = fifth
ws['L2'].value = sixth
ws['N2'].value = seventh

# Save workbook
wb.save(config['location']['address'])